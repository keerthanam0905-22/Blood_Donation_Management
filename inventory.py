from openpyxl import load_workbook

filename = "stock.xlsx"
LOW_STOCK_LIMIT = 5   # Alert if units <= 5


def add_units():
    blood = input("Enter Blood Group: ").upper()
    units_to_add = int(input("Enter Units to Add: "))

    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == blood:
            row[1].value += units_to_add
            workbook.save(filename)
            print("✅ Units Added Successfully!")
            return

    print("❌ Blood group not found!")


def remove_units():
    blood = input("Enter Blood Group: ").upper()
    units_to_remove = int(input("Enter Units to Remove: "))

    workbook = load_workbook(filename)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == blood:
            if row[1].value < units_to_remove:
                print("❌ Not enough stock available!")
                return

            row[1].value -= units_to_remove
            workbook.save(filename)
            print("✅ Units Removed Successfully!")

            # Low stock alert after removal
            if row[1].value <= LOW_STOCK_LIMIT:
                print("⚠ ALERT: Stock is Low for", blood)

            return

    print("❌ Blood group not found!")


def display_stock():
    workbook = load_workbook(filename)
    sheet = workbook.active

    print("\n📊 Available Blood Stock:")
    print("----------------------------")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        blood_group = row[0]
        units = row[1]
        print(f"{blood_group} : {units} units")

        if units <= LOW_STOCK_LIMIT:
            print(f"⚠ ALERT: Low stock for {blood_group}")

    print("----------------------------\n")


# Main Menu Loop
while True:
    print("===== Blood Bank Inventory System =====")
    print("1. Add Units")
    print("2. Remove Units")
    print("3. Display Stock")
    print("4. Exit")

    choice = input("Enter choice: ")

    if choice == "1":
        add_units()
    elif choice == "2":
        remove_units()
    elif choice == "3":
        display_stock()
    elif choice == "4":
        print("Exiting system...")
        break
    else:
        print("Invalid choice! Try again.")